# -*- coding: utf-8 -*-
import io
import os
import re
import random
from datetime import datetime
from typing import Dict, Tuple, List, Set

import requests
import streamlit as st
import openpyxl
from dateutil.relativedelta import relativedelta
from docx import Document
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

st.set_page_config(
    page_title="Εργαλείο Κατανομής Ανθρωπομηνών (v2 decimals)",
    page_icon="📊",
    layout="wide",
)

# ------------------------- GitHub assets -------------------------
TEMPLATE_URL = "https://raw.githubusercontent.com/dimitrisaronis1-dev/MANMONTHS-11/main/OUTPUT%20-%20NEW%20FORMAT%20TEMPLATE.xlsx"
LOGO_URL = "https://raw.githubusercontent.com/dimitrisaronis1-dev/MANMONTHS-11/main/SPACE%20LOGO_colored%20horizontal.png"
GUIDE_DOC_URL = "https://raw.githubusercontent.com/dimitrisaronis1-dev/MANMONTHS-11/main/%CE%9F%CE%B4%CE%B7%CE%B3%CE%AF%CE%B5%CF%82%20%CF%87%CF%81%CE%AE%CF%83%CE%B7%CF%82%20%CE%95%CF%81%CE%B3%CE%B1%CE%BB%CE%B5%CE%AF%CE%BF%CF%85%20%CE%9A%CE%B1%CF%84%CE%B1%CE%BD%CE%BF%CE%BC%CE%AE%CF%82%20%CE%91%CE%9C.docx"
LOGO_WIDTH_PX = 380  # ~10cm

# ------------------------- Styles -------------------------
st.markdown(
    """
<style>
.block-container { padding-top: 1.6rem; padding-bottom: 2rem; }
h1, h2, h3 { letter-spacing: -0.02em; }

.notice {
  border-left: 4px solid #4C78A8;
  background: rgba(76,120,168,0.08);
  padding: 10px 12px;
  border-radius: 10px;
}
.successbox {
  border-left: 4px solid #2E7D32;
  background: rgba(46,125,50,0.08);
  padding: 10px 12px;
  border-radius: 10px;
}
.errorbox {
  border-left: 4px solid #C62828;
  background: rgba(198,40,40,0.08);
  padding: 10px 12px;
  border-radius: 10px;
}
hr { margin: 1.2rem 0 1.2rem 0; }

/* Sticky bottom controls in sidebar */
section[data-testid="stSidebar"] > div {
  display: flex;
  flex-direction: column;
  height: 100%;
}
.sidebar-bottom {
  margin-top: auto;
  padding-top: 12px;
  padding-bottom: 6px;
}
</style>
""",
    unsafe_allow_html=True,
)

# ------------------------- Excel settings & styles -------------------------
MAX_YEARLY_CAPACITY = 11
YELLOW_RGB_CANDIDATES = {"FFFF00", "FFFFFF00"}

yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _http_get_bytes(url: str, timeout: int = 30) -> bytes:
    r = requests.get(url, timeout=timeout)
    r.raise_for_status()
    return r.content


@st.cache_data(show_spinner=False)
def get_template_bytes() -> bytes:
    return _http_get_bytes(TEMPLATE_URL)


@st.cache_data(show_spinner=False)
def get_logo_bytes() -> bytes:
    return _http_get_bytes(LOGO_URL)


@st.cache_data(show_spinner=False)
def get_guide_doc_bytes() -> bytes:
    return _http_get_bytes(GUIDE_DOC_URL)


@st.cache_data(show_spinner=False)
def get_guide_text() -> str:
    """Extract plain text from the DOCX guide."""
    doc_bytes = get_guide_doc_bytes()
    doc = Document(io.BytesIO(doc_bytes))
    paras = []
    for p in doc.paragraphs:
        t = (p.text or "").strip()
        if t:
            paras.append(t)
    return "\n\n".join(paras).strip()


def norm_period(s: str) -> str:
    return str(s).replace("—", "-").replace("–", "-").replace("  ", " ").strip()


def parse_date(text: str, is_start: bool = True) -> datetime:
    t = str(text).strip()
    if "σήμερα" in t.lower() or "simera" in t.lower():
        if not is_start:
            return datetime.today()
        raise ValueError("Το 'Σήμερα' επιτρέπεται μόνο ως ημερομηνία λήξης.")

    if re.match(r"^\d{4}$", t):
        return datetime.strptime(("01/01/" if is_start else "31/12/") + t, "%d/%m/%Y")

    if re.match(r"^\d{1,2}/\d{4}$", t):
        if is_start:
            return datetime.strptime("01/" + t, "%d/%m/%Y")
        d = datetime.strptime("01/" + t, "%d/%m/%Y")
        return d + relativedelta(months=1) - relativedelta(days=1)

    if re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", t):
        return datetime.strptime(t, "%d/%m/%Y")

    raise ValueError(f"Unsupported date format: {t}")


def parse_period(p: str) -> Tuple[datetime, datetime]:
    p = norm_period(p)
    if re.match(r"^\d{4}$", p):
        return parse_date(p, True), parse_date(p, False)
    parts = [x.strip() for x in p.split("-")]
    if len(parts) != 2:
        raise ValueError(f"Invalid period: {p}")
    return parse_date(parts[0], True), parse_date(parts[1], False)


def month_range(start: datetime, end: datetime) -> List[Tuple[int, int]]:
    cur = datetime(start.year, start.month, 1)
    endm = datetime(end.year, end.month, 1)
    out: List[Tuple[int, int]] = []
    while cur <= endm:
        out.append((cur.year, cur.month))
        cur += relativedelta(months=1)
    return out


def is_light_color(hex_color: str) -> bool:
    hex_color = hex_color.lstrip("#")
    rgb = tuple(int(hex_color[i : i + 2], 16) for i in (0, 2, 4))
    luminance = (0.299 * rgb[0] + 0.587 * rgb[1] + 0.114 * rgb[2]) / 255
    return luminance > 0.5



class Model:
    """
    v2 model (decimals):

    We allocate in STEPS of 0.1 AM.
    - 1.0 AM  == 10 steps (10 units)
    - Year cap: 11.0 AM == 110 steps
    - Month cap: 1.0 AM == 10 steps

    To keep logic close to v1, each step is treated like a tiny "X" that occupies a slot.
    A month can have up to 10 slots (possibly owned by many rows).
    """
    UNITS_PER_AM = 10
    MAX_YEARLY_CAPACITY_AM = 11.0
    MAX_MONTHLY_CAPACITY_AM = 1.0

    MAX_YEARLY_CAPACITY_STEPS = int(MAX_YEARLY_CAPACITY_AM * UNITS_PER_AM)  # 110
    MAX_MONTHLY_CAPACITY_STEPS = int(MAX_MONTHLY_CAPACITY_AM * UNITS_PER_AM)  # 10

    def __init__(self):
        self.ym_to_col: Dict[Tuple[int, int], int] = {}
        self.col_to_ym: Dict[int, Tuple[int, int]] = {}
        self.year_blocks: List[Tuple[int, int, int]] = []

        self.row_period: Dict[int, str] = {}
        self.row_months: Dict[int, Set[Tuple[int, int]]] = {}
        self.row_months_list: Dict[int, List[Tuple[int, int]]] = {}

        # requested amounts
        self.row_requested_am: Dict[int, float] = {}
        self.row_requested_steps: Dict[int, int] = {}

        self.row_is_yellow: Dict[int, bool] = {}
        self.row_project_index: Dict[int, int] = {}

        # allocation state
        self.month_slots: Dict[Tuple[int, int], List[int]] = {}  # ym -> [row,row,...] length<=10
        self.yearly_totals_steps: Dict[int, int] = {}           # year -> steps allocated
        self.row_steps_count: Dict[int, int] = {}               # row -> steps allocated

    def is_yellow_row(self, r: int) -> bool:
        return bool(self.row_is_yellow.get(r, False))

    def donor_can_deallocate(self, r: int) -> bool:
        # Rule 3: keep >= 1 step if possible
        return self.row_steps_count.get(r, 0) > 1

    def month_total_steps(self, ym: Tuple[int, int]) -> int:
        return len(self.month_slots.get(ym, []))

    def year_total_steps(self, year: int) -> int:
        return self.yearly_totals_steps.get(year, 0)

    def row_total_am(self, r: int) -> float:
        return round(self.row_steps_count.get(r, 0) / self.UNITS_PER_AM, 1)


def build_headers_and_maps(ws, years: List[int], start_col: int, year_row: int, month_row: int) -> Tuple[Dict[Tuple[int, int], int], int]:
    col = start_col
    month_col_map: Dict[Tuple[int, int], int] = {}

    for y in years:
        year_start = col
        year_cell = ws.cell(year_row, col)

        rc = lambda: random.randint(0, 255)
        rand_hex = "%02X%02X%02X" % (rc(), rc(), rc())
        year_cell.fill = PatternFill(start_color=rand_hex, end_color=rand_hex, fill_type="solid")
        year_cell.font = Font(color="FFFFFF" if not is_light_color(rand_hex) else "000000")

        for m in range(1, 13):
            ws.cell(month_row, col).value = m
            month_col_map[(y, m)] = col
            col += 1

        ws.merge_cells(start_row=year_row, start_column=year_start, end_row=year_row, end_column=col - 1)
        year_cell.value = y

    for c in range(start_col, col):
        ws.cell(year_row, c).border = thin_border
        ws.cell(month_row, c).border = thin_border

    return month_col_map, col


# Custom Excel format: show "X" when cell == 1, else show 1 decimal
MONTH_CELL_NUMBER_FORMAT = '[=1]"X";0.0'


def _cell_add_step(ws, r: int, c: int, delta_steps: int) -> None:
    """Update a month cell by +/-0.1 AM per step."""
    cell = ws.cell(r, c)
    cur = cell.value
    cur = float(cur) if cur not in (None, "") else 0.0
    new_val = round(cur + (delta_steps / Model.UNITS_PER_AM), 1)
    if new_val <= 0:
        cell.value = None
    else:
        cell.value = new_val
        cell.number_format = MONTH_CELL_NUMBER_FORMAT


def move_x(model: Model, ws, r: int, from_ym: Tuple[int, int], to_ym: Tuple[int, int]) -> bool:
    """
    v2: move ONE step (0.1 AM) from from_ym -> to_ym for the same row.
    """
    if from_ym not in model.ym_to_col or to_ym not in model.ym_to_col:
        return False
    if to_ym not in model.row_months.get(r, set()):
        return False

    # must own at least one step in from_ym
    slots = model.month_slots.get(from_ym, [])
    if r not in slots:
        return False

    # destination must have month capacity
    if model.month_total_steps(to_ym) >= Model.MAX_MONTHLY_CAPACITY_STEPS:
        return False

    # yearly capacity if changing year
    if to_ym[0] != from_ym[0] and model.year_total_steps(to_ym[0]) >= Model.MAX_YEARLY_CAPACITY_STEPS:
        return False

    # remove 1 step from source
    slots.remove(r)
    model.month_slots[from_ym] = slots
    model.yearly_totals_steps[from_ym[0]] = model.year_total_steps(from_ym[0]) - 1
    model.row_steps_count[r] = model.row_steps_count.get(r, 0) - 1
    _cell_add_step(ws, r, model.ym_to_col[from_ym], -1)

    # add 1 step to dest
    model.month_slots.setdefault(to_ym, []).append(r)
    model.yearly_totals_steps[to_ym[0]] = model.year_total_steps(to_ym[0]) + 1
    model.row_steps_count[r] = model.row_steps_count.get(r, 0) + 1
    _cell_add_step(ws, r, model.ym_to_col[to_ym], +1)
    return True


def add_x(model: Model, ws, r: int, ym: Tuple[int, int]) -> bool:
    """
    v2: add ONE step (0.1 AM) to row r at month ym.
    """
    if ym not in model.row_months.get(r, set()):
        return False
    if model.month_total_steps(ym) >= Model.MAX_MONTHLY_CAPACITY_STEPS:
        return False
    if model.year_total_steps(ym[0]) >= Model.MAX_YEARLY_CAPACITY_STEPS:
        return False

    model.month_slots.setdefault(ym, []).append(r)
    model.yearly_totals_steps[ym[0]] = model.year_total_steps(ym[0]) + 1
    model.row_steps_count[r] = model.row_steps_count.get(r, 0) + 1
    _cell_add_step(ws, r, model.ym_to_col[ym], +1)
    return True


def remove_one_step(model: Model, ws, donor_row: int, ym: Tuple[int, int]) -> bool:
    """Remove ONE step from donor_row in ym (if present)."""
    slots = model.month_slots.get(ym, [])
    if donor_row not in slots:
        return False
    slots.remove(donor_row)
    model.month_slots[ym] = slots
    model.yearly_totals_steps[ym[0]] = model.year_total_steps(ym[0]) - 1
    model.row_steps_count[donor_row] = model.row_steps_count.get(donor_row, 0) - 1
    _cell_add_step(ws, donor_row, model.ym_to_col[ym], -1)
    return True


def free_capacity_in_year(model: Model, ws, year: int) -> bool:
    """
    Try to free 1 step of yearly capacity for `year` without violating core rules.

    Strategy:
    1) Move one step from `year` to another year within SAME project's period (white first, then yellow).
    2) If moving is impossible, deallocate 1 step from a donor in `year` but ONLY if donor keeps >=1 step
       (white first; yellow only if already allowed by Rule 3 logic elsewhere).
    """
    # collect all (ym, row) step-allocations in this year
    allocs = []
    for ym, owners in model.month_slots.items():
        if ym[0] != year:
            continue
        for r in owners:
            allocs.append((ym, r))
    allocs.sort(key=lambda t: (model.is_yellow_row(t[1]), model.row_steps_count.get(t[1], 0)))

    # 1) Try to move out of year
    for allow_yellow in (False, True):
        for ym_from, r in allocs:
            if (not allow_yellow) and model.is_yellow_row(r):
                continue
            for ym_to in model.row_months_list.get(r, []):
                if ym_to[0] == year:
                    continue
                if model.month_total_steps(ym_to) >= Model.MAX_MONTHLY_CAPACITY_STEPS:
                    continue
                if model.year_total_steps(ym_to[0]) >= Model.MAX_YEARLY_CAPACITY_STEPS:
                    continue
                if move_x(model, ws, r, ym_from, ym_to):
                    return True

    # 2) Deallocate 1 step (white first), only if donor keeps >=1 step
    for allow_yellow in (False, True):
        for ym_from, r in allocs:
            if (not allow_yellow) and model.is_yellow_row(r):
                continue
            if not model.donor_can_deallocate(r):
                continue
            # do not deallocate from yellow here; keep yellow protected by default
            if model.is_yellow_row(r):
                continue
            if remove_one_step(model, ws, r, ym_from):
                return True

    return False


def make_month_free(model: Model, ws, ym: Tuple[int, int], depth: int = 0, max_depth: int = 6, visited=None) -> bool:
    """
    v2: Ensure month `ym` has at least 1 free step slot (total < 10).
    If full, try to relocate ONE step belonging to some occupant row to another allowed month.
    """
    if model.month_total_steps(ym) < Model.MAX_MONTHLY_CAPACITY_STEPS:
        return True

    if visited is None:
        visited = set()
    if (ym, depth) in visited:
        return False
    visited.add((ym, depth))

    # month is full: try moving one step from an occupant out
    occupants = list(model.month_slots.get(ym, []))
    # Prefer white donors with more steps, then yellow
    occupants.sort(key=lambda r: (model.is_yellow_row(r), -model.row_steps_count.get(r, 0)))

    for occ_r in occupants:
        from_ym = ym
        # candidate destinations: same-year months first then other years
        dests_same = [d for d in model.row_months_list.get(occ_r, []) if d != from_ym and d[0] == from_ym[0]]
        dests_other = [d for d in model.row_months_list.get(occ_r, []) if d != from_ym and d[0] != from_ym[0]]

        for to_ym in dests_same + dests_other:
            # If destination month is full, try freeing it (chain)
            if model.month_total_steps(to_ym) >= Model.MAX_MONTHLY_CAPACITY_STEPS:
                if depth >= max_depth:
                    continue
                if not make_month_free(model, ws, to_ym, depth + 1, max_depth, visited):
                    continue

            # If destination year full, try freeing capacity in that year
            if to_ym[0] != from_ym[0] and model.year_total_steps(to_ym[0]) >= Model.MAX_YEARLY_CAPACITY_STEPS:
                if not free_capacity_in_year(model, ws, to_ym[0]):
                    continue

            if move_x(model, ws, occ_r, from_ym, to_ym):
                return True

    return False


def compute_unallocated_reasons(model: Model, target_row: int) -> str:
    req_steps = model.row_requested_steps.get(target_row, 0)
    alloc_steps = model.row_steps_count.get(target_row, 0)
    if req_steps <= 0 or alloc_steps >= req_steps:
        return ""

    reasons = set()
    months = model.row_months_list.get(target_row, [])
    for (y, m) in months:
        ym = (y, m)
        if model.month_total_steps(ym) >= Model.MAX_MONTHLY_CAPACITY_STEPS:
            reasons.add(f"Month {m}/{y} is full")
        if model.year_total_steps(y) >= Model.MAX_YEARLY_CAPACITY_STEPS:
            reasons.add(f"Year {y} capacity reached")

    if not reasons:
        return "Capacity constraints."
    ordered = sorted(reasons, key=lambda s: (s.startswith("Year "), s))
    return "; ".join(ordered)


def build_summary_text(model: Model, impossible_rows: List[int]) -> str:
    lines = []
    lines.append(f"Μέγιστη ετήσια χωρητικότητα ανά έτος: {Model.MAX_YEARLY_CAPACITY_AM:.1f} ΑΜ ({Model.MAX_YEARLY_CAPACITY_STEPS} units)\n")
    lines.append("Ετήσια σύνολα ανθρωπομηνών:\n")
    for y in sorted(model.yearly_totals_steps.keys()):
        total_steps = model.yearly_totals_steps.get(y, 0)
        total_am = round(total_steps / Model.UNITS_PER_AM, 1)
        if total_steps >= Model.MAX_YEARLY_CAPACITY_STEPS:
            lines.append(f"Έτος {y}: {total_am:.1f} (Η χωρητικότητα επιτεύχθηκε)")
        else:
            lines.append(f"Έτος {y}: {total_am:.1f}")
    lines.append("")

    deficits = []
    for r in sorted(model.row_period.keys()):
        req_steps = model.row_requested_steps.get(r, 0)
        alloc_steps = model.row_steps_count.get(r, 0)
        if req_steps > alloc_steps:
            deficits.append((r, req_steps, alloc_steps, req_steps - alloc_steps))

    if deficits:
        lines.append("Έργα με μη κατανεμημένους ανθρωπομήνες:\n")
        for r, req_s, alloc_s, unalloc_s in deficits:
            period = model.row_period.get(r, "")
            req_am = round(req_s / Model.UNITS_PER_AM, 1)
            alloc_am = round(alloc_s / Model.UNITS_PER_AM, 1)
            unalloc_am = round(unalloc_s / Model.UNITS_PER_AM, 1)
            lines.append(f"Περίοδος: {period}, Αρχικοί ΑΜ: {req_am:.1f}, Κατανεμημένοι ΑΜ: {alloc_am:.1f}, Μη κατανεμημένοι ΑΜ: {unalloc_am:.1f}")
            reasons = compute_unallocated_reasons(model, r)
            if reasons:
                lines.append(f"Λόγοι μη κατανομής: {reasons}")
            if r in impossible_rows:
                lines.append("Σημείωση: IMPOSSIBLE (δεν υπήρχε μαθηματικά διαθέσιμη χωρητικότητα μετά από όλες τις μεταφορές/steals).")
            lines.append("")
    else:
        lines.append("Δεν υπάρχουν έργα με μη κατανεμημένους ανθρωπομήνες.\n")

    return "\n".join(lines).strip()


def copy_style(style_obj):
    import copy
    return copy.copy(style_obj)


def _get_local_template_bytes() -> bytes:
    """
    v2: Prefer local OUTPUT template if available (supports offline runs),
    otherwise fall back to GitHub template URL.
    """
    candidate_paths = [
        # user-provided template in this workspace
        "/mnt/data/INPUT - NEW FORMAT TEMPLATE-2_ΚΑΤΑΝΟΜΗ ΑΜ.xlsx",
        "/mnt/data/OUTPUT - NEW FORMAT TEMPLATE.xlsx",
        os.path.join(os.getcwd(), "OUTPUT - NEW FORMAT TEMPLATE.xlsx"),
    ]
    for p in candidate_paths:
        if os.path.exists(p):
            with open(p, "rb") as f:
                return f.read()
    return get_template_bytes()


def process_excel(input_bytes: bytes, input_filename: str = "") -> Tuple[bytes, str]:
    wb_in = openpyxl.load_workbook(io.BytesIO(input_bytes))
    ws_in = wb_in.active

    headers = {str(ws_in.cell(1, c).value).strip(): c for c in range(1, ws_in.max_column + 1)}
    if "ΧΡΟΝΙΚΟ ΔΙΑΣΤΗΜΑ" not in headers or "ΑΝΘΡΩΠΟΜΗΝΕΣ" not in headers:
        raise RuntimeError("Το input πρέπει να έχει στήλες: ΧΡΟΝΙΚΟ ΔΙΑΣΤΗΜΑ και ΑΝΘΡΩΠΟΜΗΝΕΣ")

    PERIOD_COL_IN = headers["ΧΡΟΝΙΚΟ ΔΙΑΣΤΗΜΑ"]
    AM_COL_IN = headers["ΑΝΘΡΩΠΟΜΗΝΕΣ"]

    rows = []
    all_months = set()

    for r in range(2, ws_in.max_row + 1):
        period_cell = ws_in.cell(r, PERIOD_COL_IN)
        period_val = period_cell.value
        am_raw = ws_in.cell(r, AM_COL_IN).value

        # parse AM as float (expects multiples of 0.1)
        try:
            am = float(am_raw) if am_raw is not None else 0.0
        except Exception:
            am = 0.0

        if not period_val or am <= 0:
            continue

        # basic validation (locked by spec: should always be multiple of 0.1)
        am_rounded = round(am, 1)
        if abs(am - am_rounded) > 1e-9:
            raise RuntimeError(f"Μη έγκυρο AM (όχι πολλαπλάσιο του 0.1) στη γραμμή {r}: {am_raw}")
        am = am_rounded
        req_steps = int(round(am * Model.UNITS_PER_AM))

        start, end = parse_period(str(period_val))
        months = month_range(start, end)
        for ym in months:
            all_months.add(ym)

        rgb = period_cell.fill.start_color.rgb if period_cell.fill.start_color else None
        is_yellow = rgb in YELLOW_RGB_CANDIDATES

        rows.append({
            "period_str": str(period_val).strip(),
            "requested_am": am,
            "requested_steps": req_steps,
            "months_set": set(months),
            "months_list": sorted(months),
            "is_yellow": is_yellow,
        })

    if not rows:
        raise RuntimeError("Δεν βρέθηκαν έγκυρες γραμμές (με διάστημα και ΑΜ > 0).")

    years = sorted(set(y for y, _ in all_months))
    rows.sort(key=lambda x: (not x["is_yellow"], len(x["months_list"])))

    template_bytes = _get_local_template_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = wb.active
    ws.freeze_panes = "F6"

    ws.title = "ΑΝΑΛΥΣΗ"

    # Ensure CV sheet exists (avoid CV1) and copy input into it
    for sname in list(wb.sheetnames):
        if re.fullmatch(r"CV\d+", sname):
            del wb[sname]

    if "CV" in wb.sheetnames:
        cv_sheet = wb["CV"]
        for row in cv_sheet.iter_rows():
            for cell in row:
                cell.value = None
                cell.fill = PatternFill()
                cell.border = Border()
                cell.font = Font()
                cell.number_format = "General"
    else:
        cv_sheet = wb.create_sheet(title="CV", index=0)

    for row_idx, row_data in enumerate(ws_in.iter_rows()):
        for col_idx, cell in enumerate(row_data):
            new_cell = cv_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell.value)
            if cell.has_style:
                new_cell.font = copy_style(cell.font)
                new_cell.border = copy_style(cell.border)
                new_cell.fill = copy_style(cell.fill)
                new_cell.number_format = cell.number_format

    for col_idx in range(1, ws_in.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if col_letter in ws_in.column_dimensions:
            cv_sheet.column_dimensions[col_letter].width = ws_in.column_dimensions[col_letter].width

    cv_sheet["A1"] = "Α/Α"
    cv_sheet["A1"].font = Font(bold=True)
    cv_sheet["A1"].border = thin_border

    last_row_b = 0
    for rr in range(1, cv_sheet.max_row + 1):
        if cv_sheet.cell(rr, 3).value is not None:
            last_row_b = rr

    for i in range(2, last_row_b + 1):
        cv_sheet.cell(i, 1).value = i - 1
        cv_sheet.cell(i, 1).border = thin_border

    START_ROW_DATA = 6
    YEAR_ROW = 2
    MONTH_ROW = 3
    YEARLY_TOTAL_ROW = 5
    START_COL = 6  # F

    for rng in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(str(rng))
        if (min_row <= YEAR_ROW <= max_row) or (min_row <= MONTH_ROW <= max_row) or (min_row <= START_ROW_DATA <= max_row) or (min_row <= YEARLY_TOTAL_ROW <= max_row):
            ws.unmerge_cells(str(rng))

    max_col_to_clear = max(START_COL + len(years) * 12, ws.max_column + 1)
    for r_clear in [YEAR_ROW, MONTH_ROW, START_ROW_DATA, YEARLY_TOTAL_ROW]:
        for c_clear in range(1, max_col_to_clear):
            ws.cell(r_clear, c_clear).value = None
            ws.cell(r_clear, c_clear).fill = PatternFill()

    for r_clear in range(START_ROW_DATA, ws.max_row + 1):
        for c_clear in range(1, max_col_to_clear):
            ws.cell(r_clear, c_clear).value = None
            ws.cell(r_clear, c_clear).fill = PatternFill()

    month_col_map, end_col = build_headers_and_maps(ws, years, START_COL, YEAR_ROW, MONTH_ROW)

    ws.cell(YEARLY_TOTAL_ROW, 2).value = "ΕΤΗΣΙΑ ΣΥΝΟΛΑ"
    ws.cell(YEARLY_TOTAL_ROW, 2).font = Font(bold=True)
    ws.cell(YEARLY_TOTAL_ROW, 2).border = thin_border

    ws["A5"] = "Α/Α"
    ws["B5"] = "ΕΡΓΑ"
    ws["C5"] = "ΧΡΟΝΙΚΟ ΔΙΑΣΤΗΜΑ"
    ws["D5"] = "ΑΜ"
    ws["E5"] = "X"

    # Title block A2:B4
    try:
        ws.unmerge_cells("A2:B4")
    except Exception:
        pass
    ws.merge_cells("A2:B4")
    title_cell = ws["A2"]
    title_text = "ΚΑΤΑΝΟΜΗ ΑΜ"
    if input_filename:
        base_name = os.path.splitext(input_filename)[0]
        title_text = f"{title_text}\n{base_name}"
    title_cell.value = title_text
    title_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_cell.fill = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")  # light blue
    title_cell.font = Font(color="FF0000", bold=True, size=11)
    for rr in range(2, 5):
        for cc in range(1, 3):
            ws.cell(rr, cc).border = thin_border

    # Experience header (as in v1)
    ws["C2"] = "ΑΜ ΕΜΠΕΙΡΙΑΣ"
    ws["C2"].fill = orange_fill
    ws["C2"].font = Font(bold=True)
    ws["C2"].border = thin_border

    # D2/E2 reserved for yellow totals (filled later)
    ws["D2"].fill = orange_fill
    ws["D2"].font = Font(bold=True)
    ws["D2"].border = thin_border
    ws["E2"].fill = orange_fill
    ws["E2"].font = Font(bold=True)
    ws["E2"].border = thin_border


    # Apply borders on header row 5
    for c in range(1, end_col):
        ws.cell(5, c).border = thin_border
        ws.cell(5, c).font = Font(bold=True)

    last_month_col_letter = openpyxl.utils.get_column_letter(end_col - 1)

    model = Model()
    for ym, c in month_col_map.items():
        model.ym_to_col[ym] = c
        model.col_to_ym[c] = ym
    for y in years:
        model.year_blocks.append((y, month_col_map[(y, 1)], month_col_map[(y, 12)]))

    current_row = START_ROW_DATA
    for idx, row in enumerate(rows, start=1):
        r = current_row
        model.row_period[r] = row["period_str"]
        model.row_months[r] = row["months_set"]
        model.row_months_list[r] = row["months_list"]
        model.row_requested_am[r] = float(row["requested_am"])
        model.row_requested_steps[r] = int(row["requested_steps"])
        model.row_is_yellow[r] = bool(row["is_yellow"])
        model.row_project_index[r] = idx
        model.row_steps_count[r] = 0

        ws.cell(r, 1).value = idx
        ws.cell(r, 1).border = thin_border

        ws.cell(r, 2).value = f"=VLOOKUP(A{r},CV!$A$2:$B$1899,2,FALSE)"
        ws.cell(r, 2).border = thin_border

        ws.cell(r, 3).value = f"=VLOOKUP(A{r},CV!$A$2:$C$1899,3,FALSE)"
        ws.cell(r, 3).border = thin_border
        if model.row_is_yellow[r]:
            ws.cell(r, 3).fill = yellow_fill

        ws.cell(r, 4).value = model.row_requested_am[r]
        ws.cell(r, 4).border = thin_border
        ws.cell(r, 4).number_format = "0.0"

        # Column E "X" becomes SUM of month allocations (numeric)
        ws.cell(r, 5).value = round(model.row_steps_count.get(r, 0) / Model.UNITS_PER_AM, 1)
        ws.cell(r, 5).border = thin_border
        ws.cell(r, 5).number_format = "0.0"

        # Red highlight on requested AM (col D) only if allocated < requested (step-accurate)
        req_s = model.row_requested_steps.get(r, 0)
        alloc_s = model.row_steps_count.get(r, 0)
        if req_s > 0 and alloc_s < req_s:
            ws.cell(r, 4).font = Font(color="FF0000", bold=True)
        else:
            ws.cell(r, 4).font = Font(bold=True)


        # mark allowed months area
        for ym in model.row_months_list[r]:
            cell = ws.cell(r, model.ym_to_col[ym])
            cell.fill = yellow_fill
            cell.border = thin_border
            cell.number_format = MONTH_CELL_NUMBER_FORMAT

        # Initial greedy allocation in steps (0.1)
        need = model.row_requested_steps[r]
        got = 0
        for ym in model.row_months_list[r]:
            if got >= need:
                break
            # fill this month as much as possible
            while got < need and model.month_total_steps(ym) < Model.MAX_MONTHLY_CAPACITY_STEPS and model.year_total_steps(ym[0]) < Model.MAX_YEARLY_CAPACITY_STEPS:
                if not add_x(model, ws, r, ym):
                    break
                got += 1

        for c in range(START_COL, end_col):
            ws.cell(r, c).border = thin_border
            ws.cell(r, c).number_format = MONTH_CELL_NUMBER_FORMAT
        current_row += 1

    impossible_rows: List[int] = []

    # Rule 3: ensure >= 1 step if possible
    for r in list(model.row_period.keys()):
        if model.row_requested_steps.get(r, 0) > 0 and model.row_steps_count.get(r, 0) == 0:
            satisfied = False
            for ym in model.row_months_list.get(r, []):
                if model.month_total_steps(ym) >= Model.MAX_MONTHLY_CAPACITY_STEPS:
                    if not make_month_free(model, ws, ym, depth=0, max_depth=6, visited=set()):
                        continue
                if model.year_total_steps(ym[0]) >= Model.MAX_YEARLY_CAPACITY_STEPS:
                    if not free_capacity_in_year(model, ws, ym[0]):
                        continue
                if add_x(model, ws, r, ym):
                    satisfied = True
                    break
            if not satisfied:
                impossible_rows.append(r)

    def optimize_rows(target_rows: List[int]) -> None:
        changed = True
        iters = 0
        while changed and iters < 12000:
            iters += 1
            changed = False
            deficits = [rr for rr in target_rows if model.row_requested_steps.get(rr, 0) > 0 and model.row_steps_count.get(rr, 0) < model.row_requested_steps.get(rr, 0)]
            if not deficits:
                break
            deficits.sort(key=lambda rr: (-(model.row_requested_steps[rr] - model.row_steps_count[rr]), len(model.row_months_list.get(rr, []))))
            for rr in deficits:
                for ym in model.row_months_list.get(rr, []):
                    if model.row_steps_count.get(rr, 0) >= model.row_requested_steps.get(rr, 0):
                        break
                    if model.month_total_steps(ym) >= Model.MAX_MONTHLY_CAPACITY_STEPS:
                        if not make_month_free(model, ws, ym, depth=0, max_depth=6, visited=set()):
                            continue
                    if model.year_total_steps(ym[0]) >= Model.MAX_YEARLY_CAPACITY_STEPS:
                        if not free_capacity_in_year(model, ws, ym[0]):
                            continue
                    if add_x(model, ws, rr, ym):
                        changed = True
                        break
                if changed:
                    break

    yellow_rows = [r for r in model.row_period if model.is_yellow_row(r)]
    white_rows = [r for r in model.row_period if not model.is_yellow_row(r)]
    optimize_rows(yellow_rows)
    optimize_rows(white_rows)

    # Lock yellow after their own optimization
    locked_yellow_steps = {r: model.row_steps_count.get(r, 0) for r in yellow_rows}

    def yellow_ok() -> bool:
        for r in yellow_rows:
            if model.row_steps_count.get(r, 0) < locked_yellow_steps.get(r, 0):
                return False
        return True

    # Improve white deficits by controlled stealing from WHITE donors (>1 step)
    def improve_white_deficits_by_steal(max_iters: int = 8000) -> None:
        it = 0
        while it < max_iters:
            it += 1
            deficits = [rr for rr in white_rows if model.row_requested_steps.get(rr, 0) > model.row_steps_count.get(rr, 0)]
            if not deficits:
                break
            deficits.sort(key=lambda rr: (-(model.row_requested_steps[rr] - model.row_steps_count[rr]), len(model.row_months_list.get(rr, []))))
            progress = False

            for rr in deficits:
                for ym in model.row_months_list.get(rr, []):
                    if model.row_steps_count.get(rr, 0) >= model.row_requested_steps.get(rr, 0):
                        break

                    # If month has room, just add
                    if model.month_total_steps(ym) < Model.MAX_MONTHLY_CAPACITY_STEPS and model.year_total_steps(ym[0]) < Model.MAX_YEARLY_CAPACITY_STEPS:
                        if add_x(model, ws, rr, ym):
                            progress = True
                            break
                        continue

                    # If month full, try steal 1 step from a WHITE donor with >1 step in same month
                    if model.month_total_steps(ym) >= Model.MAX_MONTHLY_CAPACITY_STEPS:
                        owners = list(model.month_slots.get(ym, []))
                        owners.sort(key=lambda r: (model.is_yellow_row(r), model.row_steps_count.get(r, 0)))  # prefer white + small?
                        stolen = False
                        for donor in owners:
                            if donor == rr:
                                continue
                            if model.is_yellow_row(donor):
                                continue
                            if not model.donor_can_deallocate(donor):
                                continue
                            # steal: remove from donor, add to rr (month total stays 10)
                            if remove_one_step(model, ws, donor, ym) and add_x(model, ws, rr, ym):
                                stolen = True
                                progress = True
                                break
                        if stolen:
                            break

                if progress:
                    break

            if not progress:
                break

    improve_white_deficits_by_steal()

    # Final hard pass: Rule 3 again (allow last-resort yellow stealing ONLY to satisfy min 1 step)
    for r in list(model.row_period.keys()):
        if model.row_requested_steps.get(r, 0) > 0 and model.row_steps_count.get(r, 0) == 0:
            satisfied = False
            for ym in model.row_months_list.get(r, []):
                # if month has room, try add
                if model.month_total_steps(ym) < Model.MAX_MONTHLY_CAPACITY_STEPS and model.year_total_steps(ym[0]) < Model.MAX_YEARLY_CAPACITY_STEPS:
                    if add_x(model, ws, r, ym):
                        satisfied = True
                        break

                # month full: steal 1 step from any donor with >1 (white first, then yellow)
                if model.month_total_steps(ym) >= Model.MAX_MONTHLY_CAPACITY_STEPS:
                    owners = list(model.month_slots.get(ym, []))
                    owners.sort(key=lambda rr: (model.is_yellow_row(rr), model.row_steps_count.get(rr, 0)))  # white first
                    for donor in owners:
                        if donor == r:
                            continue
                        if not model.donor_can_deallocate(donor):
                            continue
                        # if donor is yellow, it's last resort: allow but try to keep yellow lock if possible
                        if model.is_yellow_row(donor):
                            if remove_one_step(model, ws, donor, ym) and add_x(model, ws, r, ym):
                                satisfied = True
                                break
                        else:
                            if remove_one_step(model, ws, donor, ym) and add_x(model, ws, r, ym):
                                satisfied = True
                                break
                    if satisfied:
                        break

            if not satisfied and r not in impossible_rows:
                impossible_rows.append(r)


    # Update column E ("X") totals per row AFTER all allocations/optimizations
    for r in list(model.row_period.keys()):
        alloc_steps = model.row_steps_count.get(r, 0)
        ws.cell(r, 5).value = round(alloc_steps / Model.UNITS_PER_AM, 1)
        ws.cell(r, 5).number_format = "0.0"
        ws.cell(r, 5).border = thin_border
        ws.cell(r, 5).font = Font(bold=True)

        # Red highlight on requested AM (col D) only if allocated < requested (step-accurate)
        req_s = model.row_requested_steps.get(r, 0)
        if req_s > 0 and alloc_steps < req_s:
            ws.cell(r, 4).font = Font(color="FF0000", bold=True)
        else:
            ws.cell(r, 4).font = Font(bold=True)

    # Yearly totals row
    ws.cell(YEARLY_TOTAL_ROW, 4).value = "Σύνολο ΑΜ"
    ws.cell(YEARLY_TOTAL_ROW, 4).font = Font(bold=True)

    for y in years:
        year_start_col = month_col_map[(y, 1)]
        year_end_col = month_col_map[(y, 12)]
        # put total in first month col cell of totals row (same as v1 pattern)
        cell = ws.cell(YEARLY_TOTAL_ROW, year_start_col)
        cell.value = round(model.year_total_steps(y) / Model.UNITS_PER_AM, 1)
        cell.number_format = "0.0"
        cell.border = thin_border
        # merge across year block for display
        try:
            ws.merge_cells(start_row=YEARLY_TOTAL_ROW, start_column=year_start_col, end_row=YEARLY_TOTAL_ROW, end_column=year_end_col)
        except Exception:
            pass
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if model.year_total_steps(y) >= Model.MAX_YEARLY_CAPACITY_STEPS:
            cell.fill = orange_fill

    # Yellow totals in D2/E2 (requested / allocated), as in v1
    yellow_req_steps = sum(model.row_requested_steps.get(r, 0) for r in model.row_period if model.is_yellow_row(r))
    yellow_alloc_steps = sum(model.row_steps_count.get(r, 0) for r in model.row_period if model.is_yellow_row(r))
    ws["D2"].value = round(yellow_req_steps / Model.UNITS_PER_AM, 1)
    ws["E2"].value = round(yellow_alloc_steps / Model.UNITS_PER_AM, 1)
    ws["D2"].number_format = "0.0"
    ws["E2"].number_format = "0.0"

    summary_text = build_summary_text(model, impossible_rows)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), summary_text

# ------------------------- Sidebar: Usage Guide (sticky bottom-left) -------------------------
with st.sidebar:
    st.markdown('<div class="sidebar-bottom">', unsafe_allow_html=True)
    st.link_button("📘 Οδηγίες χρήσης της εφαρμογής", "?page=guide", use_container_width=True)
    st.download_button(
        label="⬇️ Λήψη οδηγιών (Word)",
        data=get_guide_doc_bytes(),
        file_name="Οδηγίες χρήσης Εργαλείου Κατανομής ΑΜ.docx",
        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        use_container_width=True,
    )
    st.markdown("</div>", unsafe_allow_html=True)


# ------------------------- Routing: main vs guide view -------------------------
page = st.query_params.get("page", "")
if page == "guide":
    left, right = st.columns([0.72, 0.28], vertical_alignment="center")
    with left:
        st.markdown("## Εργαλείο Κατανομής Ανθρωπομηνών")
        st.markdown("### Οδηγίες χρήσης της εφαρμογής")
    with right:
        try:
            st.image(get_logo_bytes(), width=LOGO_WIDTH_PX)
        except Exception:
            pass

    st.markdown("---")
    try:
        guide_text = get_guide_text()
        if guide_text:
            st.markdown(guide_text.replace("\n", "  \n"))
        else:
            st.info("Δεν βρέθηκε κείμενο στο αρχείο οδηγιών.")
    except Exception as e:
        st.error(f"Αποτυχία φόρτωσης οδηγιών: {e}")

    st.markdown("---")
    st.link_button("⬅️ Επιστροφή στην εφαρμογή", "?", use_container_width=False)
    st.stop()


# ------------------------- Main page -------------------------
left, right = st.columns([0.72, 0.28], vertical_alignment="center")
with left:
    st.markdown("## Εργαλείο Κατανομής Ανθρωπομηνών_v2")
    st.markdown("Αυτό το εργαλείο κατανέμει ανθρωπομήνες σε έργα με βάση χρονικά διαστήματα και μέγιστη ετήσια χωρητικότητα.")
with right:
    try:
        st.image(get_logo_bytes(), width=LOGO_WIDTH_PX)
    except Exception:
        pass

st.markdown("---")
st.markdown("### 👉 Ανέβασε το INPUT excel (μόνο 3 στήλες)")

uploaded = st.file_uploader(
    " ",
    type=["xlsx"],
    accept_multiple_files=False,
    help="Το Excel πρέπει να έχει στήλες: ΧΡΟΝΙΚΟ ΔΙΑΣΤΗΜΑ και ΑΝΘΡΩΠΟΜΗΝΕΣ.",
)

if not uploaded:
    st.markdown('<div class="notice">Παρακαλώ ανεβάστε ένα αρχείο Excel για να ξεκινήσετε την επεξεργασία.</div>', unsafe_allow_html=True)
    st.stop()

st.write(f"**Επιλεγμένο αρχείο:** `{uploaded.name}`")
run_btn = st.button("✅ Εκτέλεση κατανομής", use_container_width=True)

if run_btn:
    with st.spinner("Εκτελείται η κατανομή..."):
        try:
            out_bytes, summary_text = process_excel(uploaded.getvalue(), input_filename=uploaded.name)
            st.session_state["out_bytes"] = out_bytes
            st.session_state["summary_text"] = summary_text
            st.session_state["out_name"] = os.path.splitext(uploaded.name)[0] + "_ΚΑΤΑΝΟΜΗ ΑΜ.xlsx"
            st.markdown('<div class="successbox">Το αρχείο επεξεργάστηκε επιτυχώς!</div>', unsafe_allow_html=True)
        except Exception as e:
            st.session_state.pop("out_bytes", None)
            st.session_state.pop("summary_text", None)
            st.markdown(f'<div class="errorbox">Αποτυχία επεξεργασίας: {e}</div>', unsafe_allow_html=True)

if "out_bytes" in st.session_state:
    st.download_button(
        label="⬇️ Κατεβάστε το επεξεργασμένο Excel",
        data=st.session_state["out_bytes"],
        file_name=st.session_state.get("out_name", "output.xlsx"),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.markdown("### **Σύνοψη Κατανομής**")
    st.text(st.session_state.get("summary_text", ""))
